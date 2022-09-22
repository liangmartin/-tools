# -*- coding:UTF-8 -*-
#-----author:ljf -----#
#----date:20220921----#
#----version:1.0------#
import datetime
from bisect import bisect
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Font
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

time = datetime.datetime.now().strftime("%Y%m%d")

_LIST1, _LIST2  = [], []
_INIT = False
dest_filename = "ip_report.xlsx"
data = {}

#判断是否是有效IP
ip2int = lambda ip_str: reduce(lambda a, b: (a << 8) + b, [int(i) for i in ip_str.split('.')])  

#判断是否是合法IP
is_ip=re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')

def _init():  
    global _LIST, _INIT  
    if not _INIT:  
        for l in open('comm/ip.merge.txt', 'rb'):  
            ip1, ip2 = l.split('|')[:2]
            addr = '|'.join(l.split('|')[2:])
            ip1, ip2 = ip2int(ip1), ip2int(ip2)  
            _LIST1.append(ip1)  
            _LIST2.append((ip1, ip2, addr))
        _INIT = True
        

def ip_from(ip):
    with pool_sema:  
        if is_ip.match(ip):
            i = ip2int(ip)
            idx = bisect(_LIST1, i)
            assert(idx > 0)
            if len(_LIST1) < idx:
                data[ip]=(u'未知 未知IP地址 %s' % ip)
            else:
                frm, to ,addr = _LIST2[idx - 1]
                if frm <= i <= to:
                    data[ip]=addr
                    #print(frm,to,addr,i)
                else:
                    data[ip]=(u'未知 未知IP地址 %s' % ip)
        else:
            data[ip]=(u'unknown %s IP不合法' % ip)
            



def creat_xlsx(data):
    wb = Workbook()
    ws1 = wb.worksheets[0]
    ws1.title = u"IP列表报告情况"
    ## set list
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 9
    ws1.column_dimensions['C'].width = 21
    ws1.column_dimensions['D'].width = 11
    ws1.column_dimensions['E'].width = 24.88
    ## edit head
    ws1['A1'] = "IP列表"
    ws1['B1'] = "IP国家"
    ws1['C1'] = "IP省份"
    ws1['D1'] = "IP城市"
    ws1['E1'] = "IP运营商"
    ##edit font
    font = Font(size=14)
    ws1['A1'].font = font
    ws1['B1'].font = font
    ws1['C1'].font = font
    ws1['D1'].font = font
    ws1['E1'].font = font
    # write data
    rows = 2
    for ip,addr in data.items():
        aa = addr.replace("0","").replace("\n","")
        ws1.cell(column=1, row=rows, value=ip)
        ws1.cell(column=2, row=rows, value=aa.split('|')[0])
        ws1.cell(column=3, row=rows, value=aa.split('|')[2])
        ws1.cell(column=4, row=rows, value=aa.split('|')[3])
        ws1.cell(column=5, row=rows, value=aa.split('|')[4])
        rows = rows + 1
    wb.save(filename=dest_filename)


#发邮件
def send_email(file_name):
    SENDTIME = (datetime.datetime.now().strftime("%Y-%m-%d %H:00:00"))
    msg = MIMEMultipart('related')
    msg.attach(MIMEText(SENDTIME + " ip_check脚本执行情况详情 请见附件，谢谢！！", "plain", "utf-8"))

    def accessory(name):
        i = MIMEText(open(name, "rb").read(), "base64", "utf-8")
        i["Content-Type"] = "application/octet-stream"
        i["Content-Disposition"] = "attachment; filename=\"{}\"".format(name)
        return i

    msg.attach(accessory(file_name))

    HEAD = u"IP报告情况"
    SUBJECT = SENDTIME + HEAD
    TO = "xxxx"
    FROM = "xxxxx"
    msg['Subject'] = SUBJECT
    msg['From'] = FROM
    msg['To'] = TO
    HOST = "xxxx" 
    server = smtplib.SMTP()
    server.connect(HOST, "25")
    server.login("xxxxxx", "xx") 
    server.sendmail(FROM, TO.split(","), msg.as_string())
    server.quit()






if __name__ == '__main__':
    max_threads=200
    pool_sema = threading.BoundedSemaphore(max_threads)
    try:
        _init()
        with open('ip.txt','rb') as f:
                threads = []
                for line in f.readlines():
                    ip = line.strip()
                    threads.append(threading.Thread(target=ip_from,args=(ip,)))
                for t in threads:
                    t.start()
                for d in threads:
                    d.join() 
        creat_xlsx(data)
        #send_email(dest_filename)              
    except Exception:
        print 'Throw an exception....'

        
